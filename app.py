from flask import Flask, request, jsonify, send_from_directory, render_template, session, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import os
import uuid
from config import Config
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)

# Load configuration
app.config.from_object(Config)

# 圖片上傳配置
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# 確保上傳資料夾存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'categories'), exist_ok=True)
os.makedirs(os.path.join(UPLOAD_FOLDER, 'products'), exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# JWT 配置已經在 config.py 中設置

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)
CORS(app)
jwt = JWTManager(app)

# JWT 錯誤處理
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    print(f"Token expired: {jwt_payload}")  # 調試信息
    return jsonify({'error': 'Token has expired'}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error):
    print(f"Invalid token: {error}")  # 調試信息
    return jsonify({'error': 'Invalid token'}), 401

@jwt.unauthorized_loader
def missing_token_callback(error):
    print(f"Missing token: {error}")  # 調試信息
    return jsonify({'error': 'Authorization token is required'}), 401

# Database Models
class Admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    description = db.Column(db.Text)
    image_url = db.Column(db.String(200))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship
    items = db.relationship('Item', backref='category', lazy=True)

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.Float, nullable=False)
    quantity_left = db.Column(db.Integer, nullable=False, default=0)
    image_url = db.Column(db.String(200))
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True)  # 是否上架
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    city = db.Column(db.String(50))
    postal_code = db.Column(db.String(10))
    country = db.Column(db.String(50), default='TW')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 為了向後兼容，保留name屬性
    @property
    def name(self):
        return f"{self.first_name} {self.last_name}".strip()
    
    @name.setter
    def name(self, value):
        parts = value.split(' ', 1)
        self.first_name = parts[0]
        self.last_name = parts[1] if len(parts) > 1 else ''

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, completed, cancelled
    order_date = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 7-11門市相關欄位
    delivery_method = db.Column(db.String(50))  # 取貨方式
    store_id = db.Column(db.String(50))  # 7-11門市代碼
    store_name = db.Column(db.String(100))  # 7-11門市名稱
    store_address = db.Column(db.Text)  # 7-11門市地址
    
    # 付款和備註
    payment_method = db.Column(db.String(50))  # 付款方式
    notes = db.Column(db.Text)  # 備註
    
    # Relationships
    customer = db.relationship('Customer', backref=db.backref('orders', lazy=True))
    order_items = db.relationship('OrderItem', backref='order', lazy=True, cascade='all, delete-orphan')

class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price_at_time = db.Column(db.Float, nullable=False)  # Store price at time of purchase
    
    # Relationships
    item = db.relationship('Item', backref=db.backref('order_items', lazy=True))

class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    store_code = db.Column(db.String(20), unique=True, nullable=False)  # 門市代碼
    store_name = db.Column(db.String(100), nullable=False)  # 門市名稱
    address = db.Column(db.String(200), nullable=False)  # 門市地址
    city = db.Column(db.String(50), nullable=False)  # 城市
    district = db.Column(db.String(50), nullable=False)  # 區域
    phone = db.Column(db.String(20))  # 門市電話
    is_active = db.Column(db.Boolean, default=True)  # 是否營業中
    
    def to_dict(self):
        return {
            'id': self.id,
            'store_code': self.store_code,
            'store_name': self.store_name,
            'address': self.address,
            'city': self.city,
            'district': self.district,
            'phone': self.phone
        }

# API Routes

# Admin Authentication
@app.route('/api/admin/register', methods=['POST'])
def admin_register():
    data = request.get_json()
    
    if not data or not data.get('username') or not data.get('email') or not data.get('password'):
        return jsonify({'error': 'Missing required fields'}), 400
    
    if Admin.query.filter_by(username=data['username']).first():
        return jsonify({'error': 'Username already exists'}), 400
    
    if Admin.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email already exists'}), 400
    
    admin = Admin(
        username=data['username'],
        email=data['email']
    )
    admin.set_password(data['password'])
    
    db.session.add(admin)
    db.session.commit()
    
    return jsonify({'message': 'Admin registered successfully'}), 201

@app.route('/api/admin/login', methods=['POST'])
def admin_login_api():
    data = request.get_json()
    
    if not data or not data.get('username') or not data.get('password'):
        return jsonify({'error': 'Missing username or password'}), 400
    
    admin = Admin.query.filter_by(username=data['username']).first()
    
    if admin and admin.check_password(data['password']):
        access_token = create_access_token(identity=str(admin.id))  # 轉換為字符串
        return jsonify({
            'access_token': access_token,
            'admin_id': admin.id,
            'username': admin.username
        }), 200
    
    return jsonify({'error': 'Invalid credentials'}), 401

# Category Management
@app.route('/api/categories', methods=['GET'])
def get_categories():
    categories = Category.query.all()
    return jsonify([{
        'id': category.id,
        'name': category.name,
        'description': category.description,
        'image_url': category.image_url,
        'created_at': category.created_at.isoformat(),
        'item_count': len([item for item in category.items if item.is_active])
    } for category in categories])

@app.route('/api/categories/<int:category_id>', methods=['GET'])
def get_category(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        return jsonify({'error': 'Category not found'}), 404
    
    return jsonify({
        'id': category.id,
        'name': category.name,
        'description': category.description,
        'image_url': category.image_url,
        'created_at': category.created_at.isoformat(),
        'item_count': len([item for item in category.items if item.is_active])
    })

@app.route('/api/categories/<int:category_id>/items', methods=['GET'])
def get_category_items(category_id):
    category = db.session.get(Category, category_id)
    if not category:
        return jsonify({'error': 'Category not found'}), 404
    
    # 只返回上架的商品
    items = [item for item in category.items if item.is_active]
    
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'description': item.description,
        'price': item.price,
        'quantity_left': item.quantity_left,
        'image_url': item.image_url,
        'category_id': item.category_id,
        'is_active': item.is_active,
        'created_at': item.created_at.isoformat()
    } for item in items])

# 7-11門市相關API
@app.route('/api/stores', methods=['GET'])
def get_stores():
    """獲取所有7-11門市"""
    try:
        stores = Store.query.filter_by(is_active=True).order_by(Store.city, Store.district, Store.store_name).all()
        stores_data = [store.to_dict() for store in stores]
        return jsonify(stores_data)
        
    except Exception as e:
        print(f"獲取門市列表錯誤: {e}")
        return jsonify({'error': f'獲取門市列表失敗: {str(e)}'}), 500

@app.route('/api/stores/<int:store_id>', methods=['GET'])
def get_store(store_id):
    """獲取特定門市詳情"""
    try:
        store = db.session.get(Store, store_id)
        if not store:
            return jsonify({'error': '門市不存在'}), 404
        
        return jsonify(store.to_dict())
        
    except Exception as e:
        print(f"獲取門市詳情錯誤: {e}")
        return jsonify({'error': f'獲取門市詳情失敗: {str(e)}'}), 500

# 7-11門市選擇回調端點
@app.route('/cvs_callback', methods=['POST'])
def cvs_callback():
    """7-11門市選擇回調"""
    try:
        # 獲取7-11 API回傳的資料
        storeid = request.form.get('storeid')
        storename = request.form.get('storename')
        storeaddress = request.form.get('storeaddress')
        outside = request.form.get('outside')
        ship = request.form.get('ship')
        tempvar = request.form.get('TempVar')
        
        # print(f"7-11門市選擇回調資料:")
        # print(f"  storeid: {storeid}")
        # print(f"  storename: {storename}")
        # print(f"  storeaddress: {storeaddress}")
        # print(f"  outside: {outside}")
        # print(f"  ship: {ship}")
        # print(f"  TempVar: {tempvar}")
        
        # 驗證必要資料
        if not storeid or not storename or not storeaddress:
            return jsonify({'error': '門市資料不完整'}), 400
        
        # 將門市資料存儲到session中
        session['selected_store'] = {
            'storeid': storeid,
            'storename': storename,
            'storeaddress': storeaddress,
            'outside': outside,
            'ship': ship,
            'tempvar': tempvar
        }
        
        # 返回成功頁面，自動跳轉回結帳頁面
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>門市選擇成功</title>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    text-align: center;
                    padding: 50px;
                    background: #f5f5f5;
                }}
                .success-box {{
                    background: white;
                    padding: 30px;
                    border-radius: 10px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    max-width: 500px;
                    margin: 0 auto;
                }}
                .success-icon {{
                    color: #28a745;
                    font-size: 48px;
                    margin-bottom: 20px;
                }}
                .store-info {{
                    background: #f8f9fa;
                    padding: 20px;
                    border-radius: 5px;
                    margin: 20px 0;
                    text-align: left;
                }}
                .btn {{
                    background: #007bff;
                    color: white;
                    padding: 10px 20px;
                    border: none;
                    border-radius: 5px;
                    cursor: pointer;
                    text-decoration: none;
                    display: inline-block;
                }}
            </style>
        </head>
        <body>
            <div class="success-box">
                <div class="success-icon">✓</div>
                <h2>門市選擇成功！</h2>
                <div class="store-info">
                    <h3>{storename}</h3>
                    <p><strong>門市代碼：</strong>{storeid}</p>
                    <p><strong>門市地址：</strong>{storeaddress}</p>
                </div>
                <p>正在跳轉回結帳頁面...</p>
                <a href="/checkout?from_store_selection=true" class="btn">返回結帳頁面</a>
            </div>
            <script>
                // 3秒後自動跳轉
                setTimeout(function() {{
                    window.location.href = '/checkout?from_store_selection=true';
                }}, 3000);
            </script>
        </body>
        </html>
        """
        
    except Exception as e:
        print(f"7-11門市選擇回調錯誤: {e}")
        return jsonify({'error': f'處理門市選擇失敗: {str(e)}'}), 500

@app.route('/api/selected-store', methods=['GET'])
def get_selected_store():
    """獲取當前選中的門市"""
    try:
        selected_store = session.get('selected_store')
        if not selected_store:
            return jsonify({'error': '尚未選擇門市'}), 404
        
        return jsonify(selected_store)
        
    except Exception as e:
        print(f"獲取選中門市錯誤: {e}")
        return jsonify({'error': f'獲取選中門市失敗: {str(e)}'}), 500

@app.route('/api/selected-store', methods=['DELETE'])
def clear_selected_store():
    """清除選中的門市"""
    try:
        session.pop('selected_store', None)
        return jsonify({'message': '門市選擇已清除'})
        
    except Exception as e:
        print(f"清除選中門市錯誤: {e}")
        return jsonify({'error': f'清除選中門市失敗: {str(e)}'}), 500

@app.route('/api/categories', methods=['POST'])
@jwt_required()
def create_category():
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    data = request.get_json()
    
    if not data or not data.get('name'):
        return jsonify({'error': 'Category name is required'}), 400
    
    if Category.query.filter_by(name=data['name']).first():
        return jsonify({'error': 'Category name already exists'}), 400
    
    category = Category(
        name=data['name'],
        description=data.get('description', ''),
        image_url=data.get('image_url', '')
    )
    
    db.session.add(category)
    db.session.commit()
    
    return jsonify({
        'message': 'Category created successfully',
        'category': {
            'id': category.id,
            'name': category.name,
            'description': category.description
        }
    }), 201

@app.route('/api/categories/<int:category_id>', methods=['PUT'])
@jwt_required()
def update_category(category_id):
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    category = db.session.get(Category, category_id)
    if not category:
        return jsonify({'error': 'Category not found'}), 404
    
    data = request.get_json()
    
    if data.get('name'):
        # Check if new name already exists
        existing_category = Category.query.filter_by(name=data['name']).first()
        if existing_category and existing_category.id != category_id:
            return jsonify({'error': 'Category name already exists'}), 400
        category.name = data['name']
    
    if data.get('description') is not None:
        category.description = data['description']
    
    db.session.commit()
    
    return jsonify({'message': 'Category updated successfully'})

@app.route('/api/categories/<int:category_id>', methods=['DELETE'])
@jwt_required()
def delete_category(category_id):
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    category = db.session.get(Category, category_id)
    if not category:
        return jsonify({'error': 'Category not found'}), 404
    
    # Check if category has items
    if category.items:
        return jsonify({'error': 'Cannot delete category with existing items'}), 400
    
    db.session.delete(category)
    db.session.commit()
    
    return jsonify({'message': 'Category deleted successfully'})

# Item Management
@app.route('/api/items', methods=['GET'])
def get_items():
    category_id = request.args.get('category_id', type=int)
    
    if category_id:
        items = Item.query.filter_by(category_id=category_id).all()
    else:
        items = Item.query.all()
    
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'description': item.description,
        'price': item.price,
        'quantity_left': item.quantity_left,
        'image_url': item.image_url,
        'is_active': item.is_active,
        'category': {
            'id': item.category.id,
            'name': item.category.name
        },
        'created_at': item.created_at.isoformat()
    } for item in items])

@app.route('/api/items', methods=['POST'])
@jwt_required()
def create_item():
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    data = request.get_json()
    
    if not data or not data.get('name') or not data.get('price') or not data.get('category_id'):
        return jsonify({'error': 'Missing required fields (name, price, category_id)'}), 400
    
    # Check if category exists
    category = db.session.get(Category, data['category_id'])
    if not category:
        return jsonify({'error': 'Category not found'}), 404
    
    item = Item(
        name=data['name'],
        description=data.get('description', ''),
        price=float(data['price']),
        quantity_left=int(data.get('quantity_left', 0)),
        image_url=data.get('image_url', ''),
        category_id=data['category_id'],
        is_active=data.get('is_active', True)
    )
    
    db.session.add(item)
    db.session.commit()
    
    return jsonify({
        'message': 'Item created successfully',
        'item': {
            'id': item.id,
            'name': item.name,
            'description': item.description,
            'price': item.price,
            'quantity_left': item.quantity_left,
            'image_url': item.image_url,
            'category': {
                'id': item.category.id,
                'name': item.category.name
            }
        }
    }), 201

@app.route('/api/items/<int:item_id>', methods=['PUT'])
@jwt_required()
def update_item(item_id):
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    item = db.session.get(Item, item_id)
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    data = request.get_json()
    
    if data.get('name'):
        item.name = data['name']
    if data.get('description') is not None:
        item.description = data['description']
    if data.get('price'):
        item.price = float(data['price'])
    if data.get('quantity_left') is not None:
        item.quantity_left = int(data['quantity_left'])
    if data.get('image_url') is not None:
        item.image_url = data['image_url']
    if data.get('category_id'):
        # Check if new category exists
        category = db.session.get(Category, data['category_id'])
        if not category:
            return jsonify({'error': 'Category not found'}), 404
        item.category_id = data['category_id']
    if data.get('is_active') is not None:
        item.is_active = bool(data['is_active'])
    
    item.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({'message': 'Item updated successfully'})

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
@jwt_required()
def delete_item(item_id):
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    item = db.session.get(Item, item_id)
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    db.session.delete(item)
    db.session.commit()
    
    return jsonify({'message': 'Item deleted successfully'})

# Shopping History
@app.route('/api/orders', methods=['POST'])
def create_order():
    """創建新訂單"""
    try:
        data = request.get_json()
        print(data)
        
        # 驗證必要欄位
        required_fields = ['customer', 'items', 'payment_method']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'缺少必要欄位: {field}'}), 400
        
        customer_data = data['customer']
        items_data = data['items']
        payment_method = data['payment_method']
        notes = data.get('notes', '')
        
        # 驗證客戶資料
        customer_required = ['first_name', 'last_name', 'email', 'phone', 'address', 'city']
        for field in customer_required:
            if field not in customer_data or not customer_data[field]:
                return jsonify({'error': f'缺少客戶資料: {field}'}), 400
        
        # 驗證商品資料
        if not items_data:
            return jsonify({'error': '訂單必須包含至少一個商品'}), 400
        
        # 檢查商品是否存在且有庫存
        total_amount = 0
        order_items = []

        for item_data in items_data:
            item_id = item_data.get('item_id')
            quantity = item_data.get('quantity', 1)
            price = item_data.get('price')
            
            if not item_id or not quantity or not price:
                return jsonify({'error': '商品資料不完整'}), 400
            
            # 查詢商品
            item = db.session.get(Item, item_id)
            if not item:
                return jsonify({'error': f'商品不存在: {item_id}'}), 404
            
            if not item.is_active:
                return jsonify({'error': f'商品已下架: {item.name}'}), 400
            
            if item.quantity_left < quantity:
                return jsonify({'error': f'庫存不足: {item.name} (庫存: {item.quantity_left})'}), 400
            
            total_amount += price * quantity
            order_items.append({
                'item': item,
                'quantity': quantity,
                'price': price
            })
        
        # 檢查客戶是否已存在（根據email）
        customer = Customer.query.filter_by(email=customer_data['email']).first()
        
        if customer:
            # 客戶已存在，更新客戶資訊
            print(f"客戶已存在，更新客戶資訊: {customer.email}")
            customer.first_name = customer_data['first_name']
            customer.last_name = customer_data['last_name']
            customer.phone = customer_data['phone']
            customer.address = customer_data['address']
            customer.city = customer_data['city']
            customer.postal_code = customer_data.get('postal_code', '')
            customer.country = customer_data.get('country', 'TW')
        else:
            # 創建新客戶記錄
            print(f"創建新客戶: {customer_data['email']}")
            customer = Customer(
                first_name=customer_data['first_name'],
                last_name=customer_data['last_name'],
                email=customer_data['email'],
                phone=customer_data['phone'],
                address=customer_data['address'],
                city=customer_data['city'],
                postal_code=customer_data.get('postal_code', ''),
                country=customer_data.get('country', 'TW')
            )
            db.session.add(customer)
        
        db.session.flush()  # 獲取客戶 ID
        
        # 獲取7-11門市資訊
        delivery_method = data.get('delivery_method', '')
        store_info = data.get('store_info', {})
        
        # 創建訂單
        order = Order(
            customer_id=customer.id,
            total_amount=total_amount,
            status='pending',
            delivery_method=delivery_method,
            store_id=store_info.get('storeid') if store_info else None,
            store_name=store_info.get('storename') if store_info else None,
            store_address=store_info.get('storeaddress') if store_info else None,
            payment_method=payment_method,
            notes=notes
        )
        db.session.add(order)
        db.session.flush()  # 獲取訂單 ID
        
        # 創建訂單項目並更新庫存
        for order_item_data in order_items:
            item = order_item_data['item']
            quantity = order_item_data['quantity']
            price = order_item_data['price']
            
            # 創建訂單項目
            order_item = OrderItem(
                order_id=order.id,
                item_id=item.id,
                quantity=quantity,
                price_at_time=price
            )
            db.session.add(order_item)
            
            # 更新庫存
            item.quantity_left -= quantity
        
        db.session.commit()
        
        return jsonify({
            'message': '訂單創建成功',
            'order_id': order.id,
            'total_amount': total_amount
        }), 201
        
    except Exception as e:
        db.session.rollback()
        print(f"創建訂單錯誤: {e}")
        return jsonify({'error': f'創建訂單失敗: {str(e)}'}), 500

@app.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order(order_id):
    """獲取單個訂單詳情"""
    try:
        order = db.session.get(Order, order_id)
        if not order:
            return jsonify({'error': '訂單不存在'}), 404
        
        # 獲取訂單項目
        order_items = []
        for order_item in order.order_items:
            order_items.append({
                'id': order_item.id,
                'item_name': order_item.item.name,
                'quantity': order_item.quantity,
                'price': order_item.price_at_time
            })
        
        return jsonify({
            'id': order.id,
            'order_date': order.order_date.isoformat(),
            'total_amount': order.total_amount,
            'status': order.status,
            'delivery_method': order.delivery_method,
            'store_id': order.store_id,
            'store_name': order.store_name,
            'store_address': order.store_address,
            'payment_method': order.payment_method,
            'notes': order.notes,
            'customer': {
                'first_name': order.customer.first_name,
                'last_name': order.customer.last_name,
                'email': order.customer.email,
                'phone': order.customer.phone,
                'address': order.customer.address,
                'city': order.customer.city,
                'postal_code': order.customer.postal_code,
                'country': order.customer.country
            },
            'items': order_items
        })
        
    except Exception as e:
        print(f"獲取訂單錯誤: {e}")
        return jsonify({'error': f'獲取訂單失敗: {str(e)}'}), 500

@app.route('/api/orders', methods=['GET'])
@jwt_required()
def get_all_orders():
    try:
        admin_id = get_jwt_identity()
        admin = db.session.get(Admin, int(admin_id))

        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        # 獲取查詢參數
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 限制每頁最大數量
        per_page = min(per_page, 100)
        
        # 構建查詢
        query = Order.query
        
        # 日期篩選
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Order.order_date >= start_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                # 包含結束日期的整天
                end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                query = query.filter(Order.order_date <= end_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400
        
        # 排序和分頁
        query = query.order_by(Order.order_date.desc())
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        # 構建響應數據
        orders_data = []
        for order in pagination.items:
            try:
                order_data = {
                    'id': order.id,
                    'total_amount': order.total_amount,
                    'status': order.status,
                    'order_date': order.order_date.isoformat(),
                    'customer_name': f"{order.customer.first_name} {order.customer.last_name}".strip() if order.customer else 'Unknown',
                    'customer_email': order.customer.email if order.customer else '',
                    'items_count': len(order.order_items) if order.order_items else 0,
                    'delivery_method': order.delivery_method,
                    'store_name': order.store_name,
                    'store_id': order.store_id,
                    'payment_method': order.payment_method
                }
                orders_data.append(order_data)
            except Exception as e:
                print(f"跳過有問題的訂單 {order.id}: {e}")
                continue
        
        return jsonify({
            'orders': orders_data,
            'pagination': {
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
    except Exception as e:
        print(f"訂單 API 錯誤: {e}")
        return jsonify({'error': f'獲取訂單失敗: {str(e)}'}), 500

@app.route('/api/orders/<int:order_id>', methods=['GET'])
@jwt_required()
def get_order_details(order_id):
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    order = db.session.get(Order, order_id)
    if not order:
        return jsonify({'error': 'Order not found'}), 404
    
    return jsonify({
        'id': order.id,
        'customer': {
            'id': order.customer.id,
            'first_name': order.customer.first_name,
            'last_name': order.customer.last_name,
            'email': order.customer.email,
            'phone': order.customer.phone,
            'address': order.customer.address,
            'city': order.customer.city,
            'postal_code': order.customer.postal_code,
            'country': order.customer.country
        },
        'total_amount': order.total_amount,
        'status': order.status,
        'order_date': order.order_date.isoformat(),
        'delivery_method': order.delivery_method,
        'store_id': order.store_id,
        'store_name': order.store_name,
        'store_address': order.store_address,
        'payment_method': order.payment_method,
        'notes': order.notes,
        'items': [{
            'id': item.id,
            'item_name': item.item.name,
            'quantity': item.quantity,
            'price': item.price_at_time
        } for item in order.order_items]
    })

@app.route('/api/orders/<int:order_id>/status', methods=['PUT'])
@jwt_required()
def update_order_status(order_id):
    """更新訂單狀態"""
    try:
        admin_id = get_jwt_identity()
        admin = db.session.get(Admin, int(admin_id))
        
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        order = db.session.get(Order, order_id)
        if not order:
            return jsonify({'error': 'Order not found'}), 404
        
        data = request.get_json()
        new_status = data.get('status')
        
        if not new_status:
            return jsonify({'error': 'Status is required'}), 400
        
        # 驗證狀態值
        valid_statuses = ['pending', 'confirmed', 'shipped', 'delivered', 'cancelled']
        if new_status not in valid_statuses:
            return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400
        
        # 更新狀態
        old_status = order.status
        order.status = new_status
        db.session.commit()
        
        return jsonify({
            'message': 'Order status updated successfully',
            'order_id': order.id,
            'old_status': old_status,
            'new_status': new_status
        }), 200
        
    except Exception as e:
        db.session.rollback()
        print(f"更新訂單狀態錯誤: {e}")
        return jsonify({'error': f'更新訂單狀態失敗: {str(e)}'}), 500

@app.route('/api/orders/search', methods=['GET'])
@jwt_required()
def search_orders():
    """搜尋訂單（支援訂單號和狀態搜尋）"""
    try:
        admin_id = get_jwt_identity()
        admin = db.session.get(Admin, int(admin_id))
        
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        # 獲取查詢參數
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        order_id = request.args.get('order_id')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # 限制每頁最大數量
        per_page = min(per_page, 100)
        
        # 構建查詢
        query = Order.query
        
        # 訂單號篩選
        if order_id:
            try:
                order_id_int = int(order_id)
                query = query.filter(Order.id == order_id_int)
            except ValueError:
                return jsonify({'error': 'Invalid order_id format'}), 400
        
        # 狀態篩選
        if status:
            valid_statuses = ['pending', 'confirmed', 'shipped', 'delivered', 'cancelled']
            if status not in valid_statuses:
                return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400
            query = query.filter(Order.status == status)
        
        # 日期篩選
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Order.order_date >= start_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                query = query.filter(Order.order_date <= end_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400
        
        # 排序和分頁
        query = query.order_by(Order.order_date.desc())
        pagination = query.paginate(
            page=page, 
            per_page=per_page, 
            error_out=False
        )
        
        orders = []
        for order in pagination.items:
            orders.append({
                'id': order.id,
                'customer_name': f"{order.customer.first_name} {order.customer.last_name}".strip(),
                'customer_email': order.customer.email,
                'total_amount': order.total_amount,
                'status': order.status,
                'order_date': order.order_date.isoformat(),
                'items_count': len(order.order_items),
                'delivery_method': order.delivery_method,
                'store_name': order.store_name,
                'store_id': order.store_id,
                'payment_method': order.payment_method
            })
        
        return jsonify({
            'orders': orders,
            'pagination': {
                'page': pagination.page,
                'per_page': pagination.per_page,
                'total': pagination.total,
                'pages': pagination.pages,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev
            }
        })
        
    except Exception as e:
        print(f"搜尋訂單錯誤: {e}")
        return jsonify({'error': f'搜尋訂單失敗: {str(e)}'}), 500

@app.route('/api/orders/export', methods=['GET'])
@jwt_required()
def export_orders_to_excel():
    """導出訂單到Excel文件"""
    try:
        admin_id = get_jwt_identity()
        admin = db.session.get(Admin, int(admin_id))
        
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        # 獲取查詢參數
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        status = request.args.get('status')
        
        # 構建查詢
        query = Order.query
        
        # 日期篩選
        if start_date:
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(Order.order_date >= start_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        
        if end_date:
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = end_datetime.replace(hour=23, minute=59, second=59)
                query = query.filter(Order.order_date <= end_datetime)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400
        
        # 狀態篩選
        if status:
            valid_statuses = ['pending', 'confirmed', 'shipped', 'delivered', 'cancelled']
            if status not in valid_statuses:
                return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400
            query = query.filter(Order.status == status)
        
        # 排序
        query = query.order_by(Order.order_date.desc())
        orders = query.all()
        
        if not orders:
            return jsonify({'error': '沒有找到符合條件的訂單'}), 404
        
        # 創建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "訂單匯出"
        
        # 設置標題行
        headers = [
            '訂單編號', '訂單日期', '客戶姓名', '客戶信箱', '客戶電話', 
            '客戶地址', '總金額', '訂單狀態', '取貨方式', '門市名稱', 
            '門市地址', '付款方式', '備註', '商品清單'
        ]
        
        # 設置標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入標題行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 寫入訂單數據
        for row, order in enumerate(orders, 2):
            # 獲取客戶信息
            customer_name = f"{order.customer.first_name} {order.customer.last_name}".strip() if order.customer else 'Unknown'
            customer_email = order.customer.email if order.customer else ''
            customer_phone = order.customer.phone if order.customer else ''
            customer_address = order.customer.address if order.customer else ''
            
            # 獲取商品清單
            items_list = []
            for order_item in order.order_items:
                items_list.append(f"{order_item.item.name} x{order_item.quantity} (${order_item.price_at_time})")
            items_text = "; ".join(items_list)
            
            # 狀態中文對照
            status_map = {
                'pending': '待處理',
                'confirmed': '已確認', 
                'shipped': '已出貨',
                'delivered': '已送達',
                'cancelled': '已取消'
            }
            status_text = status_map.get(order.status, order.status)
            
            # 取貨方式中文對照
            delivery_map = {
                '711_store': '7-11店到店',
                'home_delivery': '宅配',
                'pickup': '自取'
            }
            delivery_text = delivery_map.get(order.delivery_method, order.delivery_method or '')
            
            # 付款方式中文對照
            payment_map = {
                'cash_on_delivery': '貨到付款',
                'credit_card': '信用卡',
                'bank_transfer': '銀行轉帳'
            }
            payment_text = payment_map.get(order.payment_method, order.payment_method or '')
            
            # 寫入數據行
            row_data = [
                order.id,
                order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
                customer_name,
                customer_email,
                customer_phone,
                customer_address,
                order.total_amount,
                status_text,
                delivery_text,
                order.store_name or '',
                order.store_address or '',
                payment_text,
                order.notes or '',
                items_text
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 自動調整列寬
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, len(orders) + 2):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # 設置最小寬度為10，最大寬度為50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 創建響應
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        if start_date and end_date:
            filename = f"orders_export_{start_date}_to_{end_date}_{date_str}.xlsx"
        elif start_date:
            filename = f"orders_export_{start_date}_from_{date_str}.xlsx"
        elif end_date:
            filename = f"orders_export_before_{end_date}_{date_str}.xlsx"
        else:
            filename = f"orders_export_all_{date_str}.xlsx"
        
        # 使用URL編碼處理中文文件名
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        
        return response
        
    except Exception as e:
        print(f"導出Excel錯誤: {e}")
        return jsonify({'error': f'導出Excel失敗: {str(e)}'}), 500


# 圖片上傳端點
@app.route('/api/upload/category', methods=['POST'])
@jwt_required()
def upload_category_image():
    try:
        admin_id = get_jwt_identity()
        admin = db.session.get(Admin, admin_id)
        
        if not admin:
            return jsonify({'error': 'Admin not found'}), 404
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            # 生成唯一檔名
            unique_filename = f"{uuid.uuid4()}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'categories', unique_filename)
            file.save(filepath)
            
            # 返回相對 URL
            image_url = f"/uploads/categories/{unique_filename}"
            return jsonify({'image_url': image_url}), 200
        
        return jsonify({'error': 'Invalid file type'}), 400
    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/api/upload/product', methods=['POST'])
@jwt_required()
def upload_product_image():
    admin_id = get_jwt_identity()
    admin = db.session.get(Admin, int(admin_id))
    
    if not admin:
        return jsonify({'error': 'Admin not found'}), 404
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        # 生成唯一檔名
        unique_filename = f"{uuid.uuid4()}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'products', unique_filename)
        file.save(filepath)
        
        # 返回相對 URL
        image_url = f"/uploads/products/{unique_filename}"
        return jsonify({'image_url': image_url}), 200
    
    return jsonify({'error': 'Invalid file type'}), 400

# 靜態檔案服務
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 提供 HTML 頁面
# 前端頁面路由
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/category/<int:category_id>')
def category_page(category_id):
    return render_template('category.html')

@app.route('/cart')
def cart_page():
    return render_template('cart.html')

@app.route('/checkout')
def checkout_page():
    return render_template('checkout.html')

@app.route('/order-success/<int:order_id>')
def order_success_page(order_id):
    return render_template('order_success.html')

# 後台管理頁面路由
@app.route('/admin')
def admin_login_page():
    return send_from_directory('templates', 'admin_login.html')

@app.route('/admin/dashboard')
def admin_dashboard_page():
    return send_from_directory('templates', 'admin_dashboard.html')

# Health check endpoint
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'message': 'Shopping API is running'})

@app.route('/api/debug/jwt', methods=['GET'])
def debug_jwt():
    """調試 JWT 配置"""
    return jsonify({
        'jwt_secret_key_set': bool(app.config.get('JWT_SECRET_KEY')),
        'jwt_expires': app.config.get('JWT_ACCESS_TOKEN_EXPIRES'),
        'secret_key_set': bool(app.config.get('SECRET_KEY'))
    }), 200

@app.route('/api/debug/test-jwt', methods=['GET'])
@jwt_required()
def debug_test_jwt():
    """測試 JWT 驗證"""
    admin_id = get_jwt_identity()
    return jsonify({
        'message': 'JWT 驗證成功',
        'admin_id': admin_id
    }), 200

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # 檢查是否有SSL證書文件
    ssl_cert = 'cert.pem'
    ssl_key = 'key.pem'
    
    if os.path.exists(ssl_cert) and os.path.exists(ssl_key):
        print("使用HTTPS模式啟動...")
        app.run(debug=True, host='0.0.0.0', port=5000, ssl_context=(ssl_cert, ssl_key))
    else:
        print("使用HTTP模式啟動...")
        print("注意：7-11 API需要HTTPS，建議使用ngrok或配置SSL證書")
        app.run(debug=True, host='0.0.0.0', port=5000)
